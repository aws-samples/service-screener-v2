import json
from utils.CustomPage.CustomPageBuilder import CustomPageBuilder
from utils.Config import Config
import openpyxl
import constants as _C

class FindingsPageBuilder(CustomPageBuilder):
    SHEETS_TO_SKIP = ['Info', 'Appendix']
    
    FRAMEWORK_JS_LIB = [
        "/jquery.dataTables.min.js",
        "-bs4/js/dataTables.bootstrap4.min.js",
        "-responsive/js/dataTables.responsive.min.js",
        "-responsive/js/responsive.bootstrap4.min.js",
        "-buttons/js/dataTables.buttons.min.js",
        "-buttons/js/buttons.bootstrap4.min.js",
        "-buttons/js/buttons.html5.min.js",
        "-buttons/js/buttons.print.min.js",
        "-buttons/js/buttons.colVis.min.js"
    ]
    FRAMEWORK_CSS_LIB = [
        "bs4/css/dataTables.bootstrap4.min.css",
        "responsive/css/responsive.bootstrap4.min.css",
        "buttons/css/buttons.bootstrap4.min.css"
    ]
    
    def customPageInit(self):
        self.wb = openpyxl.load_workbook(_C.ROOT_DIR + '/' + Config.get('HTML_ACCOUNT_FOLDER_PATH') + '/workItem.xlsx')
        self.initCSS()
        self.initJSLib()
        return
    
    def initJSLib(self):
        pref = '../res/plugins/datatables'
        for js in self.FRAMEWORK_JS_LIB:
            self.addJSLib(pref + js)
            
    def initCSS(self):
        pref = "../res/plugins/datatables-"
        for css in self.FRAMEWORK_CSS_LIB:
            self.addCSSLib(pref + css)
    
    def getSheetTitle(self):
        columnTitles = ['Service']
        wb = self.wb
        for sheetName in wb.sheetnames:
            if sheetName in self.SHEETS_TO_SKIP:
                continue
            

            ws = wb[sheetName]
            for i in range(1, ws.max_column +1):
                columnTitles.append(ws.cell(row=1, column=i).value)
            return columnTitles
        
    def genTableHTML(self, table_id='findings-table', filter_suppressed=None):
        """
        Generate table HTML for findings
        
        Args:
            table_id: ID for the table element
            filter_suppressed: None (all), True (suppressed only), False (non-suppressed only)
        """
        wb = self.wb
        columnTitles = self.getSheetTitle()
        
        tableHTMLList = [
            f"<table id='{table_id}' class='table table-bordered table-striped'>",
            "<thead><tr>"
        ]

        if not columnTitles:
            return ''

        for title in columnTitles:
            tableHTMLList.append("<th>" + title + "</th>")
        tableHTMLList.append("</tr></thead>")
        
        tableHTMLList.append("<tbody>")
        for sheetName in wb.sheetnames:
            if sheetName not in self.SHEETS_TO_SKIP:
                ws = wb[sheetName]
                for i in range(2, ws.max_row + 1):
                    # Get the status column (last column) to filter by suppression
                    status_col_idx = ws.max_column
                    status_value = ws.cell(row=i, column=status_col_idx).value
                    
                    # Apply filter based on suppression status
                    if filter_suppressed is not None:
                        is_suppressed = status_value == 'Suppressed'
                        if filter_suppressed and not is_suppressed:
                            continue  # Skip non-suppressed when showing suppressed only
                        elif not filter_suppressed and is_suppressed:
                            continue  # Skip suppressed when showing non-suppressed only
                    
                    tableHTMLList.append("<tr><td>" + sheetName + "</td>")
                    for j in range (1, ws.max_column + 1):
                        tableHTMLList.append("<td>" + ws.cell(row=i, column=j).value + "</td>")
                    tableHTMLList.append("</tr>")
        
        tableHTMLList.append("</tbody></table>")
        
        return ''.join(tableHTMLList)
        
    
    def buildContentSummary_customPage(self):
        output = []
        
        # Generate the tabbed interface using AdminLTE nav-tabs
        tabsHTML = f'''
        <div class="card card-primary card-tabs">
            <div class="card-header p-0 pt-1">
                <ul class="nav nav-tabs" id="findings-tabs" role="tablist">
                    <li class="nav-item">
                        <a class="nav-link active" id="findings-tab" data-toggle="pill" href="#findings" role="tab" aria-controls="findings" aria-selected="true">
                            Findings
                        </a>
                    </li>
                    <li class="nav-item">
                        <a class="nav-link" id="suppressed-tab" data-toggle="pill" href="#suppressed" role="tab" aria-controls="suppressed" aria-selected="false">
                            Suppressed
                        </a>
                    </li>
                </ul>
            </div>
            <div class="card-body">
                <div class="tab-content" id="findings-tabContent">
                    <div class="tab-pane fade show active" id="findings" role="tabpanel" aria-labelledby="findings-tab">
                        {self.genTableHTML('findings-table', filter_suppressed=False)}
                    </div>
                    <div class="tab-pane fade" id="suppressed" role="tabpanel" aria-labelledby="suppressed-tab">
                        {self.genTableHTML('suppressed-table', filter_suppressed=True)}
                    </div>
                </div>
            </div>
        </div>
        '''
        
        items = [[tabsHTML, '']]
        output.append(self.generateRowWithCol(12, items, "data-context='settingTable'"))
        
        return output
    
    def buildContentDetail_customPage(self):
        js = '''
// Initialize DataTable for Findings tab
var findingsTable = $("#findings-table").DataTable({
    "responsive": true, 
    "lengthChange": true, 
    "autoWidth": false,
    "pageLength": 50,
    "buttons": ["copy", "csv", "colvis"]
});
findingsTable.buttons().container().appendTo('#findings-table_wrapper .col-md-6:eq(0)');

// Initialize DataTable for Suppressed tab
var suppressedTable = $("#suppressed-table").DataTable({
    "responsive": true, 
    "lengthChange": true, 
    "autoWidth": false,
    "pageLength": 50,
    "buttons": ["copy", "csv", "colvis"]
});
suppressedTable.buttons().container().appendTo('#suppressed-table_wrapper .col-md-6:eq(0)');

// Handle hash-based filtering
var hash = window.location.hash.slice(1);
if (hash){
    findingsTable.search(decodeURI(hash)).draw();
    suppressedTable.search(decodeURI(hash)).draw();
}

// Handle tab switching and redraw tables for proper responsive behavior
$('a[data-toggle="pill"]').on('shown.bs.tab', function (e) {
    var target = $(e.target).attr("href");
    if (target === '#findings') {
        findingsTable.columns.adjust().responsive.recalc();
    } else if (target === '#suppressed') {
        suppressedTable.columns.adjust().responsive.recalc();
    }
});
'''

        self.addJS(js)
        return
    