import os
import subprocess
import json
import shutil
from utils.Config import Config
from utils.Tools import _info, _warn
from services.PageBuilder import PageBuilder
from services.dashboard.DashboardPageBuilder import DashboardPageBuilder
from frameworks.FrameworkPageBuilder import FrameworkPageBuilder
import constants as _C


class OutputGenerator:
    """
    Orchestrates the generation of both AdminLTE (legacy) and Cloudscape (React) outputs.
    
    When beta_mode=True, generates both UIs for comparison.
    When beta_mode=False, generates only AdminLTE (legacy) for backward compatibility.
    """
    
    def __init__(self, beta_mode=False):
        """
        Initialize OutputGenerator.
        
        Args:
            beta_mode (bool): If True, generate both AdminLTE and Cloudscape.
                             If False, generate only AdminLTE (legacy).
        """
        self.beta_mode = beta_mode
        self.contexts = None
        self.regions = None
        self.account_id = None
        self.html_folder = None
        self.frameworks = []
        
    def generate(self, contexts, regions, frameworks=None):
        """
        Generate output based on beta_mode.
        
        Args:
            contexts (dict): Service scan results
            regions (list): List of AWS regions scanned
            frameworks (list): List of compliance frameworks to generate
        """
        self.contexts = contexts
        self.regions = regions
        self.frameworks = frameworks or []
        
        # Get account info
        sts_info = Config.get('stsInfo')
        self.account_id = sts_info['Account']
        self.html_folder = Config.get('HTML_ACCOUNT_FOLDER_FULLPATH')
        
        # Always generate legacy HTML (for backward compatibility)
        _info("Generating AdminLTE HTML output...")
        self._generate_legacy()
        
        # If beta mode, also generate Cloudscape
        if self.beta_mode:
            _info("Beta mode enabled - Generating Cloudscape React UI...")
            try:
                success = self._generate_cloudscape()
                if success:
                    _info("Cloudscape UI generated successfully!")
                else:
                    _warn("Cloudscape build failed. Only AdminLTE HTML is available.")
            except Exception as e:
                _warn(f"Cloudscape generation failed: {e}. Only AdminLTE HTML is available.")
        
    def _generate_legacy(self):
        """
        Generate AdminLTE HTML using existing PageBuilder.
        This maintains backward compatibility.
        """
        from services.Reporter import Reporter
        from utils.ExcelBuilder import ExcelBuilder
        from utils.CustomPage.CustomPage import CustomPage
        
        # Get parameters
        params = []
        for key, val in Config.get('_SS_PARAMS').items():
            if val != '':
                tmp = '--' + key + ' ' + str(val)
                params.append(tmp)
        
        summary = Config.get("SCREENER-SUMMARY")
        excel_obj = ExcelBuilder(self.account_id, " ".join(params))
        
        api_result_array = {}
        
        # Generate service pages
        for service, data_sets in self.contexts.items():
            result_sets = data_sets['results']
            chart_sets = data_sets['charts']
            
            reporter = Reporter(service)
            reporter.process(result_sets).processCharts(chart_sets).getSummary().getDetails()
            
            # Get PageBuilder dynamically
            from Screener import Screener
            page_builder_class = Screener.getServicePagebuilderDynamically(service)
            pb = page_builder_class(service, reporter)
            pb.buildPage()
            
            # Generate Excel
            if service not in ['guardduty']:
                suppressed_card_summary = reporter.getSuppressedCardSummary()
                excel_obj.generateWorkSheet(service, reporter.cardSummary, suppressed_card_summary)
            
            # Store API results
            if service not in api_result_array:
                api_result_array[service] = {'summary': {}, 'detail': {}, 'stats': {}}
            
            api_result_array[service]['summary'] = reporter.getCard()
            api_result_array[service]['detail'] = reporter.getDetail()
            
            # Add service statistics from stat.json file
            stat_file = os.path.join(_C.FORK_DIR, f'{service}.stat.json')
            if os.path.exists(stat_file):
                try:
                    with open(stat_file, 'r') as f:
                        stat_data = json.load(f)
                        # Add suppressed count from reporter
                        stat_data['suppressed'] = reporter.suppressedCount
                        api_result_array[service]['stats'] = stat_data
                        _info(f"Added stats for {service}: {stat_data}")
                except Exception as e:
                    _warn(f"Failed to load stats for {service}: {e}")
                    api_result_array[service]['stats'] = {
                        'resources': 0,
                        'rules': 0,
                        'exceptions': 0,
                        'timespent': 0,
                        'suppressed': 0
                    }
            else:
                api_result_array[service]['stats'] = {
                    'resources': 0,
                    'rules': 0,
                    'exceptions': 0,
                    'timespent': 0,
                    'suppressed': reporter.suppressedCount
                }
        
        # Generate dashboard
        dash_pb = DashboardPageBuilder('index', [])
        dash_pb.buildPage()
        
        # If beta mode, rename the legacy index.html to preserve it
        if self.beta_mode:
            legacy_index = os.path.join(self.html_folder, 'index.html')
            legacy_index_renamed = os.path.join(self.html_folder, 'index-legacy.html')
            if os.path.exists(legacy_index):
                os.rename(legacy_index, legacy_index_renamed)
                _info(f"Legacy index.html renamed to index-legacy.html")
        
        # Build Excel summary
        excel_obj.buildSummaryPage(summary)
        excel_obj._save()
        _info("Excel workItem.xlsx generation complete")
        
        # Now build Findings page (requires workItem.xlsx to exist)
        from utils.CustomPage.CustomPage import CustomPage
        cp = CustomPage()
        cp.buildFindingsPage()
        
        # Generate framework pages and add framework data to API
        if len(self.frameworks) > 0:
            for framework in self.frameworks:
                o = FrameworkPageBuilder(framework, api_result_array)
                if o.getGateCheckStatus():
                    o.buildPage()
                    
                    # Add framework data to api_result_array for Cloudscape UI
                    framework_key = f"framework_{framework}"
                    api_result_array[framework_key] = {
                        'metadata': o.framework.getMetaData(),
                        'summary': o.framework.generateGraphInformation(),
                        'details': o.framework.generateMappingInformation()
                    }
                    _info(f"Added framework data for {framework} to API")
                else:
                    print(framework + " GATECHECK==FALSE")
        
        # Custom pages are already built in main.py, no need to rebuild here
        # This prevents duplicate TA/COH API calls
        _info("Using CustomPage data already collected in main.py")
        
        # Add metadata including suppressions
        _info("Adding metadata to api-full.json...")
        suppression_data = self._get_suppression_data()
        api_result_array['__metadata'] = {
            'accountId': self.account_id,
            'regions': self.regions,
            'suppressions': suppression_data
        }
        _info(f"Metadata added with account: {self.account_id}, suppressions: {type(suppression_data)}")
        
        # Add CustomPage data for Cloudscape UI
        _info("Adding CustomPage data to api-full.json...")
        self._add_custompage_data(api_result_array)
        
        # Generate api-full.json
        json_path = self.html_folder + "/api-full.json"
        _info(f"Writing api-full.json to: {json_path}")
        with open(json_path, "w") as f:
            json.dump(api_result_array, f)
        _info("api-full.json written successfully")
        
        # Generate TA data for Cloudscape UI (skip if custom pages disabled)
        if Config.get('disable_custom_pages', False):
            _info("TA data generation skipped (disabled via --disable-custom-pages)")
        else:
            try:
                from Screener import Screener
                Screener.generateTAData(self.html_folder)
            except Exception as e:
                _warn(f"Failed to generate TA data: {e}")
    
    def _generate_cloudscape(self):
        """
        Build and embed React app.
        
        Returns:
            bool: True if successful, False otherwise
        """
        # Build React app
        build_result = self._build_react_app()
        
        if not build_result:
            return False
        
        # Embed data into HTML
        try:
            self._embed_data()
            return True
        except Exception as e:
            _warn(f"Failed to embed data: {e}")
            return False
    
    def _build_react_app(self):
        """
        Run npm build for React app.
        
        Returns:
            bool: True if successful, False otherwise
        """
        cloudscape_ui_dir = os.path.join(_C.ROOT_DIR, 'cloudscape-ui')
        
        # Check if cloudscape-ui directory exists
        if not os.path.exists(cloudscape_ui_dir):
            _warn(f"Cloudscape UI directory not found: {cloudscape_ui_dir}")
            return False
        
        # Check if node_modules exists, if not run npm install
        node_modules = os.path.join(cloudscape_ui_dir, 'node_modules')
        if not os.path.exists(node_modules):
            _info("Installing npm dependencies...")
            try:
                result = subprocess.run(
                    ['npm', 'install'],
                    cwd=cloudscape_ui_dir,
                    capture_output=True,
                    text=True,
                    timeout=300  # 5 minutes timeout
                )
                if result.returncode != 0:
                    _warn(f"npm install failed: {result.stderr}")
                    return False
            except subprocess.TimeoutExpired:
                _warn("npm install timed out")
                return False
            except Exception as e:
                _warn(f"npm install failed: {e}")
                return False
        
        # Run npm build
        _info("Building Cloudscape React UI (this may take 15-20 seconds)...")
        try:
            result = subprocess.run(
                ['npm', 'run', 'build'],
                cwd=cloudscape_ui_dir,
                capture_output=True,
                text=True,
                timeout=120  # 2 minutes timeout
            )
            
            if result.returncode != 0:
                _warn(f"npm build failed: {result.stderr}")
                return False
            
            _info("✓ Cloudscape React UI build completed successfully")
            return True
            
        except subprocess.TimeoutExpired:
            _warn("npm build timed out")
            return False
        except Exception as e:
            _warn(f"npm build failed: {e}")
            return False
    
    def _get_suppression_data(self):
        """
        Get suppression data from Config.
        
        Returns:
            dict: Suppression data with service-level and resource-specific suppressions
        """
        try:
            # Get suppression file from CLI params
            _cli_options = Config.get('_SS_PARAMS', {})
            suppression_file = _cli_options.get('suppress_file', None)
            
            _info(f"Looking for suppression file: {suppression_file}")
            
            if not suppression_file:
                _info("No suppression file configured")
                return {'serviceLevelSuppressions': [], 'resourceSuppressions': []}
                
            if not os.path.exists(suppression_file):
                _warn(f"Suppression file not found: {suppression_file}")
                return {'serviceLevelSuppressions': [], 'resourceSuppressions': []}
            
            # Read suppression file
            _info(f"Reading suppression file: {suppression_file}")
            with open(suppression_file, 'r') as f:
                suppression_config = json.load(f)
            
            _info(f"Loaded suppression config with keys: {suppression_config.keys()}")
            
            # Extract suppressions array
            if 'suppressions' in suppression_config:
                suppressions = suppression_config['suppressions']
                
                # Categorize suppressions
                service_level = []
                resource_specific = []
                
                for supp in suppressions:
                    service = supp.get('service', 'Unknown')
                    rule = supp.get('rule', 'Unknown')
                    resources = supp.get('resources', [])
                    resource_id = supp.get('resource_id', [])
                    reason = supp.get('reason', 'No reason provided')
                    
                    # Handle both 'resources' and 'resource_id' keys
                    all_resources = []
                    if resources:
                        all_resources.extend(resources if isinstance(resources, list) else [resources])
                    if resource_id:
                        all_resources.extend(resource_id if isinstance(resource_id, list) else [resource_id])
                    
                    if all_resources and len(all_resources) > 0:
                        # Resource-specific suppression
                        resource_specific.append({
                            'service': service,
                            'rule': rule,
                            'resources': all_resources,
                            'reason': reason
                        })
                    else:
                        # Service-level suppression
                        service_level.append({
                            'service': service,
                            'rule': rule,
                            'description': reason
                        })
                
                result = {
                    'serviceLevelSuppressions': service_level,
                    'resourceSuppressions': resource_specific
                }
                _info(f"Found {len(service_level)} service-level and {len(resource_specific)} resource-specific suppressions")
                return result
            
            _info("No 'suppressions' key found in suppression config")
            return {'serviceLevelSuppressions': [], 'resourceSuppressions': []}
            
        except Exception as e:
            _warn(f"Failed to load suppression data: {e}")
            import traceback
            traceback.print_exc()
            return {'serviceLevelSuppressions': [], 'resourceSuppressions': []}
    
    def _embed_data(self):
        """
        Embed JSON data into HTML file for offline access.
        """
        _info("Embedding scan data into Cloudscape HTML...")
        # Source files
        dist_dir = os.path.join(_C.ROOT_DIR, 'cloudscape-ui', 'dist')
        source_html = os.path.join(dist_dir, 'index.html')
        json_file = os.path.join(self.html_folder, 'api-full.json')
        
        # Destination
        dest_html = os.path.join(self.html_folder, 'index.html')
        
        # Check if source files exist
        if not os.path.exists(source_html):
            raise FileNotFoundError(f"Built HTML not found: {source_html}")
        
        if not os.path.exists(json_file):
            raise FileNotFoundError(f"JSON data not found: {json_file}")
        
        # Read HTML
        with open(source_html, 'r', encoding='utf-8') as f:
            html_content = f.read()
        
        # Read JSON
        with open(json_file, 'r', encoding='utf-8') as f:
            json_data = f.read()
        
        # Read TA data if it exists
        ta_file = os.path.join(self.html_folder, 'ta.json')
        ta_data = '{}'
        if os.path.exists(ta_file):
            with open(ta_file, 'r', encoding='utf-8') as f:
                ta_data = f.read()
        
        # Escape special characters in JSON for embedding
        # Use json.dumps to properly escape the JSON string for JavaScript
        import json
        
        # Parse and re-serialize to ensure proper escaping
        try:
            parsed_json = json.loads(json_data)
            escaped_json = json.dumps(parsed_json, separators=(',', ':'))
        except json.JSONDecodeError:
            # Fallback to manual escaping if JSON parsing fails
            escaped_json = json_data.replace('\\', '\\\\').replace('`', '\\`').replace('$', '\\$').replace('</script>', '<\\/script>')
        
        try:
            parsed_ta_data = json.loads(ta_data)
            escaped_ta_data = json.dumps(parsed_ta_data, separators=(',', ':'))
        except json.JSONDecodeError:
            # Fallback to manual escaping if JSON parsing fails
            escaped_ta_data = ta_data.replace('\\', '\\\\').replace('`', '\\`').replace('$', '\\$').replace('</script>', '<\\/script>')
        
        # Create data embedding script
        data_script = f'''<script>
window.__REPORT_DATA__ = {escaped_json};
window.__TA_DATA__ = {escaped_ta_data};
window.__ACCOUNT_ID__ = "{self.account_id}";'''
        
        # Add content enrichment data if available (Task 4.1)
        enriched_content_data = Config.get('enriched_content_data', None)
        if enriched_content_data:
            try:
                # Parse and re-serialize to ensure proper escaping
                import json
                parsed_content_data = json.loads(enriched_content_data)
                escaped_content_data = json.dumps(parsed_content_data, separators=(',', ':'))
                data_script += f'''
window.__CONTENT_ENRICHMENT_DATA__ = {escaped_content_data};'''
                _info("Content enrichment data embedded for Cloudscape UI")
            except json.JSONDecodeError as e:
                _warn(f"Failed to embed content enrichment data: {str(e)}")
        
        data_script += '''
</script>'''
        
        # Insert before closing </head> tag
        html_content = html_content.replace('</head>', f'{data_script}\n</head>')
        
        # Write to destination
        with open(dest_html, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        _info(f"✓ Data embedded successfully into Cloudscape HTML")

    def _add_custompage_data(self, api_result_array):
        """
        Add CustomPage data to api_result_array for Cloudscape UI.
        Extracts data from Excel and JSON files.
        """
        # Check if CustomPage processing is disabled
        if Config.get('disable_custom_pages', False):
            _info("CustomPage data extraction skipped (disabled via --disable-custom-pages)")
            # Set empty data structures for Cloudscape UI compatibility
            api_result_array['customPage_findings'] = {'error': 'CustomPage processing disabled', 'findings': []}
            api_result_array['customPage_modernize'] = {'error': 'CustomPage processing disabled', 'computes': {}, 'databases': {}}
            api_result_array['customPage_ta'] = {'error': 'CustomPage processing disabled', 'pillars': {}}
            api_result_array['customPage_coh'] = {'error': 'CustomPage processing disabled', 'executive_summary': {}, 'recommendations': []}
            return
        
        try:
            # Extract CPFindings from Excel
            _info("Extracting CPFindings from Excel...")
            api_result_array['customPage_findings'] = self._extract_findings_from_excel()
            
            # Extract CPModernize from JSON files
            _info("Extracting CPModernize from JSON files...")
            api_result_array['customPage_modernize'] = self._extract_modernize_from_json()
            
            # Extract CPTA from JSON files
            _info("Extracting CPTA from JSON files...")
            api_result_array['customPage_ta'] = self._extract_ta_data()
            
            # Extract COH (Cost Optimization Hub) data from JSON files
            _info("Extracting COH data from JSON files...")
            api_result_array['customPage_coh'] = self._extract_coh_data()
            
            _info("CustomPage data added successfully")
        except Exception as e:
            _warn(f"Failed to add CustomPage data: {e}")
            import traceback
            traceback.print_exc()
    
    def _extract_findings_from_excel(self):
        """
        Extract findings data from workItem.xlsx for CPFindings page.
        Returns dict with columns, findings, and suppressed items.
        """
        try:
            import openpyxl
            
            excel_path = self.html_folder + '/workItem.xlsx'
            
            if not os.path.exists(excel_path):
                _info(f"Excel file not found: {excel_path}")
                return {'columns': [], 'findings': [], 'suppressed': []}
            
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            findings = []
            suppressed = []
            columns = []
            
            sheets_to_skip = ['Info', 'Appendix']
            
            # Get column headers from first sheet
            for sheet_name in wb.sheetnames:
                if sheet_name not in sheets_to_skip:
                    ws = wb[sheet_name]
                    if ws.max_row > 0:
                        columns = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
                    break
            
            # Extract data from all sheets
            for sheet_name in wb.sheetnames:
                if sheet_name in sheets_to_skip:
                    continue
                
                ws = wb[sheet_name]
                
                # Skip if sheet is empty
                if ws.max_row < 2:
                    continue
                
                # Extract rows
                for row in range(2, ws.max_row + 1):
                    finding = {'service': sheet_name}
                    
                    # Extract all columns
                    for col_idx, header in enumerate(columns, 1):
                        cell_value = ws.cell(row, col_idx).value
                        # Convert None to empty string
                        finding[header if header else f'Column{col_idx}'] = cell_value if cell_value is not None else ''
                    
                    # Separate by status (last column is typically Status)
                    status = finding.get('Status', finding.get('status', ''))
                    if status == 'Suppressed':
                        suppressed.append(finding)
                    else:
                        findings.append(finding)
            
            _info(f"Extracted {len(findings)} findings and {len(suppressed)} suppressed items from Excel")
            
            return {
                'columns': columns,
                'findings': findings,
                'suppressed': suppressed
            }
            
        except ImportError:
            _warn("openpyxl not installed. CPFindings data will not be available.")
            return {'columns': [], 'findings': [], 'suppressed': []}
        except Exception as e:
            _warn(f"Failed to extract findings from Excel: {e}")
            import traceback
            traceback.print_exc()
            return {'columns': [], 'findings': [], 'suppressed': []}
    
    def _extract_modernize_from_json(self):
        """
        Extract modernization data from CustomPage.Modernize.*.json files.
        Returns dict with Computes and Databases Sankey diagram data.
        """
        try:
            import glob
            
            modernize_files = glob.glob(_C.FORK_DIR + '/CustomPage.Modernize.*.json')
            
            if not modernize_files:
                _info("No Modernize JSON files found")
                return {}
            
            # Import and use the Modernize class to process the data
            from utils.CustomPage.Pages.Modernize.Modernize import Modernize
            
            # Collect all service data
            all_data = {}
            for file_path in modernize_files:
                try:
                    with open(file_path, 'r') as f:
                        service_data = json.load(f)
                        # Extract service name from filename (e.g., "ec2" from "CustomPage.Modernize.ec2.json")
                        service_name = os.path.basename(file_path).split('.')[2]
                        all_data[service_name] = service_data
                        _info(f"Loaded modernize data for {service_name}")
                except Exception as e:
                    _warn(f"Failed to read {file_path}: {e}")
            
            if not all_data:
                _info("No valid modernize data found")
                return {}
            
            # Create Modernize instance and process the data
            modernize = Modernize()
            modernize.setData(all_data)  # Set the raw data
            modernize.build()  # Process into Sankey diagram data
            
            # Extract the processed Sankey data
            sankey_data = modernize.ds  # This contains {Computes: {nodes: [], links: []}, Databases: {...}}
            
            _info(f"Processed Modernize data with sections: {list(sankey_data.keys())}")
            
            return sankey_data
            
        except Exception as e:
            _warn(f"Failed to extract Modernize data: {e}")
            import traceback
            traceback.print_exc()
            return {}
    
    def _extract_ta_data(self):
        """
        Extract Trusted Advisor data from ta.json file (generated by Screener.generateTAData).
        Returns dict with TA check results by pillar.
        """
        try:
            # First try to load from the new ta.json file
            ta_json_path = os.path.join(self.html_folder, 'ta.json')
            
            if os.path.exists(ta_json_path):
                _info(f"Loading TA data from: {ta_json_path}")
                with open(ta_json_path, 'r') as f:
                    ta_data = json.load(f)
                
                _info(f"Loaded TA data with error: '{ta_data.get('error', 'None')}' and pillars: {list(ta_data.get('pillars', {}).keys())}")
                return ta_data
            
            # Fallback: try to load from legacy CustomPage.TA.*.json files
            import glob
            
            ta_files = glob.glob(_C.FORK_DIR + '/CustomPage.TA.*.json')
            
            if not ta_files:
                _info("No TA JSON files found")
                return {'error': 'No Trusted Advisor data available', 'pillars': {}}
            
            # TA data structure: {COST_OPTIMIZING: {rows: [], thead: [], total: {}}, ...}
            ta_data = {'error': '', 'pillars': {}}
            
            for file_path in ta_files:
                try:
                    with open(file_path, 'r') as f:
                        data = json.load(f)
                        # The data is already in the correct format: {"error": "", "pillars": {...}}
                        if 'error' in data and data['error']:
                            ta_data['error'] = data['error']
                        if 'pillars' in data:
                            ta_data['pillars'].update(data['pillars'])
                except Exception as e:
                    _warn(f"Failed to read {file_path}: {e}")
            
            # Check if there's an error in the TA data
            if ta_data['error']:
                _info(f"TA data contains error: {ta_data['error']}")
                return ta_data
            
            if not ta_data['pillars']:
                _info("TA data contains no pillars")
                return {'error': 'No Trusted Advisor data available', 'pillars': {}}
            
            _info(f"Extracted TA data with pillars: {list(ta_data['pillars'].keys())}")
            
            return ta_data
            
        except Exception as e:
            _warn(f"Failed to extract TA data: {e}")
            import traceback
            traceback.print_exc()
            return {'error': f'Failed to load Trusted Advisor data: {str(e)}', 'pillars': {}}

    def _extract_coh_data(self):
        """
        Extract Cost Optimization Hub data from CustomPage.COH.*.json files.
        Returns dict with COH recommendations and executive summary.
        """
        try:
            import glob
            
            coh_files = glob.glob(_C.FORK_DIR + '/CustomPage.COH.*.json')
            
            if not coh_files:
                _info("No COH JSON files found")
                return {
                    'executive_summary': {},
                    'recommendations': [],
                    'error_messages': ['No Cost Optimization Hub data available'],
                    'data_collection_time': None
                }
            
            # COH data should be aggregated from all service files
            all_coh_data = {}
            
            for file_path in coh_files:
                try:
                    with open(file_path, 'r') as f:
                        service_data = json.load(f)
                        # Extract service name from filename
                        service_name = os.path.basename(file_path).split('.')[2]
                        all_coh_data[service_name] = service_data
                        _info(f"Loaded COH data for {service_name}")
                except Exception as e:
                    _warn(f"Failed to read {file_path}: {e}")
            
            if not all_coh_data:
                _info("No valid COH data found")
                return {
                    'executive_summary': {},
                    'recommendations': [],
                    'error_messages': ['Failed to load Cost Optimization Hub data'],
                    'data_collection_time': None
                }
            
            # Import and use the COH class to get the processed data
            from utils.CustomPage.Pages.COH.COH import COH
            
            # Create COH instance and get the built data
            coh = COH()
            coh.setData(all_coh_data)  # Set the raw data
            coh.build()  # Process the data
            
            # Get the processed data for UI consumption
            coh_ui_data = coh.get_data_for_ui()
            
            _info(f"Processed COH data with {len(coh_ui_data.get('recommendations', []))} recommendations")
            
            return coh_ui_data
            
        except Exception as e:
            _warn(f"Failed to extract COH data: {e}")
            import traceback
            traceback.print_exc()
            return {
                'executive_summary': {},
                'recommendations': [],
                'error_messages': [f'Failed to load Cost Optimization Hub data: {str(e)}'],
                'data_collection_time': None
            }
