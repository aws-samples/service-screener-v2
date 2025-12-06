import boto3
import json
from datetime import datetime
from dateutil.relativedelta import relativedelta
import os
import openpyxl
import logging

## Initialization
SnsEmailSubject = "[AWS] Service Screener Scheduler Report"
SnsEmailBodyPrefix = "Here is your Service Screener report.\n\n"
snsPrefix = os.environ['SSV2_SNSARN_PREFIX']


successResp = {
    'statusCode': 200,
    'body': 'success'
}

currentResults = {}
previousResults = {}
accounts = {}

SHEETS_TO_SKIP = ['Info', 'Appendix']

# ▼ ここから追加：日本語表示用のマッピングとヘルパー類 ▼

# 重大度の日本語マッピング
SEVERITY_JA = {
    "High": "高",
    "Medium": "中",
    "Low": "低",
    "Informational": "情報",
}

# カテゴリ（Well-Architected Pillar）の日本語マッピング
PILLAR_JA = {
    "Security": "セキュリティ",
    "Cost Optimization": "コスト最適化",
    "Performance Efficiency": "パフォーマンス効率",
    "Reliability": "信頼性",
    "Operation Excellence": "運用上の優秀性",
}

# リソース種別の日本語マッピング（必要に応じて拡張）
RESOURCE_TYPE_JA = {
    "EC2": "EC2インスタンス",
    "EBS": "EBSボリューム",
    "ELB": "ロードバランサー",
    "SG": "セキュリティグループ",
    "User": "IAMユーザー",
    "Role": "IAMロール",
    "Lambda": "Lambda関数",
    "Bucket": "S3バケット",
    "mysql": "RDS（MySQL）",
}


def parse_finding_line(raw_line: str) -> dict:
    """
    1 行の検知文字列を分解して dict にする。
    例:
      'ap-northeast-1::EC2DiskMonitor::Performance Efficiency::EC2::i-xxx::Medium'
    """
    line = raw_line.strip()
    if line.startswith("--"):
        line = line[2:].strip()

    parts = line.split("::")
    if len(parts) < 6:
        # 想定外形式の場合は生文字列だけ返す
        return {"raw": raw_line}

    region, rule_name, pillar, resource_type, resource_id, severity = parts[:6]

    return {
        "region": region,
        "rule_name": rule_name,
        "pillar": pillar,
        "pillar_ja": PILLAR_JA.get(pillar, pillar),
        "resource_type": resource_type,
        "resource_type_ja": RESOURCE_TYPE_JA.get(resource_type, resource_type),
        "resource_id": resource_id,
        "severity": severity,
        "severity_ja": SEVERITY_JA.get(severity, severity),
        "raw": raw_line,
    }


def auto_generic_description(finding: dict) -> str:
    """
    カテゴリ（pillar）に応じた汎用的な日本語説明を返す。
    ルールごとの詳細辞書は作らず、「新しいルールにも柔軟に対応する」方針。
    """
    pillar = finding.get("pillar")
    if pillar == "Security":
        return "セキュリティに関する検知です。関連する設定やアクセス制御を確認してください。"
    if pillar == "Cost Optimization":
        return "コスト最適化に関する検知です。不要なリソースや設定がないか確認し、コスト削減を検討してください。"
    if pillar == "Performance Efficiency":
        return "パフォーマンス効率に関する検知です。リソースの性能やスケーリング設定を確認してください。"
    if pillar == "Reliability":
        return "信頼性に関する検知です。冗長化やバックアップ、障害時の挙動を確認してください。"
    if pillar == "Operation Excellence":
        return "運用性に関する検知です。監視・運用プロセスやオペレーションの改善を検討してください。"

    # 未分類の場合のフォールバック
    return "この検知の詳細はルール名とリソース情報を参考に確認してください。"


def render_findings_block(blocks):
    """
    compareXlsx の trow[5] / trow[6] に入っている " -- ...\n -- ..." の
    ブロック群から、日本語を交えた箇条書きリストを生成する。
    """
    results = []

    for b in blocks:
        if not b:
            continue

        for line in b.splitlines():
            line = line.strip()
            if not line:
                continue

            f = parse_finding_line(line)

            # 想定フォーマットでパースできない場合は、そのまま表示
            if "rule_name" not in f:
                results.append(f"- {f['raw'].strip()}")
                continue

            desc = auto_generic_description(f)

            # 1件分を複数行で出力（プレーンテキスト SNS メール前提）
            results.append(
                f"- [{f['severity']}/{f['severity_ja']}] {f['rule_name']}"
            )
            results.append(
                f"  カテゴリ: {f['pillar_ja']} ({f['pillar']})"
            )
            results.append(
                f"  リソース: {f['resource_type_ja']} ({f['resource_type']})"
            )
            results.append(
                f"  対象: {f['resource_id']} / リージョン: {f['region']}"
            )
            results.append(
                f"  概要: {desc}"
            )
            results.append("")  # 各検知の間を 1 行空ける

    return results

# ▲ ここまで追加ヘルパー類 ▲


def lambda_handler(event, context):
    record = event['detail']
    region = event['region']
    targetBucket = record['bucket']['name']
    targetObject = record['object']['key']
    
    s3 = boto3.client('s3', region_name=region)
    sns = boto3.client('sns', region_name=region)
    
    filepath = targetObject.split('/')
    configId = filepath[0]
    currentDate = filepath[1]

    objs = s3.list_objects_v2(Bucket=targetBucket, Prefix=configId + '/' + currentDate + '/')
    contents = objs.get('Contents')
    for xlsx in contents:
        if xlsx.get('Key').endswith('xlsx'):
            _obj = xlsx.get('Key').split('/')
            accounts[_obj[2]] = {'currentRun': currentDate}

    # find latest available filesx
    currentDateObj = datetime.strptime(currentDate, '%Y%m%d').date()
    for x in range(3):
        prefixSearch = (currentDateObj - relativedelta(months=x)).strftime("%Y%m")
        objs = s3.list_objects_v2(Bucket=targetBucket, Prefix=configId + '/' + prefixSearch)
        contents = objs.get('Contents')
        if not contents:
            continue 

        contents.reverse()
        for file in contents:
            _file = file.get('Key')
            if _file.endswith('xlsx'):
                info = _file.split('/')
                acct = info[2]
                if acct in accounts and accounts[acct].get('previousRun') == None and info[1] != currentDate:
                    accounts[acct]['previousRun'] = info[1]

        if checkIfPreviousScanFound() == True:
            break

    html = []
    for acct, info in accounts.items():
        html.append("\n AccountId: {} \n".format(acct))
        html.append( processXlsx(s3, targetBucket, configId, acct, info) )
    
    res = sendSnsEmail(sns, configId, html)
    if res[0] == False:
        return  {
            'statusCode': 500,
            'body': res[1]
        }

    return successResp

def processXlsx(s3, targetBucket, configId, acct, info):
    latestObjname = "{}/{}/{}/workItem.xlsx".format(configId, info['currentRun'], acct)
    s3.download_file(targetBucket, latestObjname, '/tmp/current.xlsx')
    
    loadXlsx('/tmp/current.xlsx')

    previousObjname = None
    hasPreviousObj = False
    if 'previousRun' in info:
        hasPreviousObj = True
        previousObjname = "{}/{}/{}/workItem.xlsx".format(configId, info['previousRun'], acct)
        s3.download_file(targetBucket, previousObjname, '/tmp/previous.xlsx')
        loadXlsx('/tmp/previous.xlsx')

    compared = compareXlsx(hasPreviousObj)
    html = formatCompared(compared, hasPreviousObj)
    
    os.remove('/tmp/current.xlsx')
    if 'previousRun' in info:
        os.remove('/tmp/previous.xlsx')

    return html

def loadXlsx(filename):
    wb = openpyxl.load_workbook(filename)
    for sheetName in wb.sheetnames:
        if sheetName in SHEETS_TO_SKIP:
            continue

        results = []
        hcnt = 0
        tcnt = 0

        ws = wb[sheetName]
        for row in ws.iter_rows(min_row=2, values_only=True):
            _row = list(row)
            if _row[0] == 'Region':
                continue

            if _row[4] == 'High':
                hcnt = hcnt + 1

            del _row[5]
            tcnt = tcnt + 1

            results.append('::'.join(_row))
            
        if filename == '/tmp/current.xlsx':
            currentResults[sheetName] = {'obj': results, 'High': hcnt, 'Total': tcnt}
        else:
            previousResults[sheetName] = {'obj': results, 'High': hcnt, 'Total': tcnt}

def compareXlsx(hasPreviousObj):
    data = []
    for sheets in currentResults:
        newFindings = []
        resolvedItems = []
        diffHigh = 0
        diffTotal = 0
        if hasPreviousObj:
            diff = list(set(currentResults[sheets]['obj']) - set(previousResults[sheets]['obj']))
            newFindings = diff
            
            diff = list(set(previousResults[sheets]['obj']) - set(currentResults[sheets]['obj']))
            resolvedItems = diff

            diffHigh = currentResults[sheets]['High'] - previousResults[sheets]['High']
            diffTotal = currentResults[sheets]['Total'] - previousResults[sheets]['Total']
        
        nf = ""
        ri = ""
        if len(newFindings):
            nf = " -- " + ("\n -- ").join(newFindings)

        if len(resolvedItems):
            ri = " -- " + ("\n -- ").join(resolvedItems)

        data.append([
            sheets,
            currentResults[sheets]['High'],
            currentResults[sheets]['Total'],
            diffHigh,
            diffTotal,
            nf,
            ri,
            len(newFindings),
            len(resolvedItems)
        ])
        logging.info(data)

    return data

def formatCompared(compared, hasPreviousObj):
    """
    Service Screener の比較結果をテキストメール用に整形する。
    サマリは表形式、New / Resolved は日本語を交えた箇条書き。
    """
    def _fmt_diff(v):
        if v > 0:
            return f"+{v}"
        if v < 0:
            return str(v)
        return "0"

    lines = []

    # ===== サマリテーブル =====
    lines.append("Summary by service")
    lines.append("------------------")

    if hasPreviousObj:
        header = "{:<15} {:>6} {:>7} {:>8} {:>10}".format(
            "SERVICE", "HIGH", "TOTAL", "ΔHIGH", "ΔTOTAL"
        )
    else:
        header = "{:<15} {:>6} {:>7}".format(
            "SERVICE", "HIGH", "TOTAL"
        )

    lines.append(header)
    lines.append("-" * len(header))

    totalNew = 0
    totalResolved = 0
    new_blocks = []
    resolved_blocks = []

    for trow in compared:
        service = trow[0]
        high = trow[1]
        total = trow[2]
        diffHigh = trow[3]
        diffTotal = trow[4]
        newStr = trow[5]
        resolvedStr = trow[6]
        newCount = trow[7]
        resolvedCount = trow[8]

        if hasPreviousObj:
            line = "{:<15} {:>6} {:>7} {:>8} {:>10}".format(
                service,
                high,
                total,
                _fmt_diff(diffHigh),
                _fmt_diff(diffTotal)
            )
            if newCount:
                totalNew += newCount
                new_blocks.append(newStr)
            if resolvedCount:
                totalResolved += resolvedCount
                resolved_blocks.append(resolvedStr)
        else:
            line = "{:<15} {:>6} {:>7}".format(
                service,
                high,
                total
            )

        lines.append(line)

    lines.append("")
    lines.append("")

    # ===== New Findings =====
    if hasPreviousObj:
        lines.append(f"New Findings (新規検知 {totalNew} 件):")
        lines.append("--------------------------------------")
        new_list = render_findings_block(new_blocks)

        if new_list:
            lines.extend(new_list)
        else:
            lines.append("  (なし)")
        lines.append("")
        lines.append("")

        # ===== Resolved =====
        lines.append(f"Resolved (解消 {totalResolved} 件):")
        lines.append("--------------------------------------")
        resolved_list = render_findings_block(resolved_blocks)

        if resolved_list:
            lines.extend(resolved_list)
        else:
            lines.append("  (なし)")
        lines.append("")

    return "\n".join(lines)


# Not in used, SNS does not support HTML unless pairing it with SES
def formatComparedHTML(compared, hasPreviousObj):

    template = "<table><thead><tr>{}</tr></thead><tbody>{}</tbody></table>"

    thead = ["Services", "# High", "# Total"]
    if hasPreviousObj:
        thead.append('# High (diff)')
        thead.append('# Total (diff)')
        thead.append("New Findings")
        thead.append("Resolved")

    theadHTML = ''.join(["<th>{}</th>".format(x) for x in thead])
    
    tbodyHTML = []
    for row in compared:
        trow = [row[0], row[1], row[2]]
        if hasPreviousObj:
            trow.append(row[3])
            trow.append(row[4])
            trow.append(row[5])
            trow.append(row[6])
        
        tbodyStr = ''.join(["<td>{}</td>".format(x) for x in trow])
        tbodyHTML.append("<tr>{}</tr>".format(tbodyStr))
    
    tbodyHTML = ''.join(tbodyHTML)
    return template.format(theadHTML, tbodyHTML)

def checkIfPreviousScanFound():
    for acct, info in accounts.items():
        if not info.get('previousRun'):
            return False

    return True
    pass

def sendSnsEmail(sns, configId, html):
    topic = snsPrefix + '-' + configId
    rrr = sns.list_topics()
    topicArn = [tp['TopicArn'] for tp in sns.list_topics()['Topics'] if topic in tp['TopicArn']]

    if len(topicArn) == 0:
        return [False, 'No SNS topic found']

    sns.publish(
        TopicArn=topicArn[0],
        Message= SnsEmailBodyPrefix + ''.join(html),
        Subject= SnsEmailSubject
    )

    return [True, 'Sent']


#with open('sampleS3Event.json', 'r') as f:
#    event = json.load(f)

# output = lambda_handler(event, '')
# print(output)
